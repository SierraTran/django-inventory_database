from unittest import TestCase
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from inventory.excel_functions import is_cell_merged, format_row, setup_worksheet

class TestExcelFunctions(TestCase):
    def setUp(self):
        """
        Setup
        """
        self.workbook = load_workbook("PO_Template.xlsx")
        self.worksheet = self.workbook.active
        
    def test_is_cell_merged(self):
        """
        Test whether the function correctly identifies merged and unmerged cells.
        """
        self.assertFalse(is_cell_merged(self.worksheet, "B2"))
        self.assertFalse(is_cell_merged(self.worksheet, "C2"))
        self.assertTrue(is_cell_merged(self.worksheet, "D2"))
        
    def test_format_row(self):
        """
        Test whether the formatting is applied correctly to a row.
        """
        row = 24
        format_row(self.worksheet, row)

        # Check border formatting
        for col in range(2, 16):  # Columns B to O
            cell = self.worksheet[f"{get_column_letter(col)}{row}"]
            self.assertIsNotNone(cell.border)

        # Check alignment formatting
        self.assertEqual(self.worksheet[f"B{row}"].alignment.vertical, "center")
        self.assertEqual(self.worksheet[f"I{row}"].alignment.vertical, "center")
        self.assertEqual(self.worksheet[f"C{row}"].alignment.horizontal, "center")

        # Check formula
        self.assertEqual(self.worksheet[f"J{row}"].value, f"=+I{row}*D{row}")
        
    def test_setup_worksheet_equal_to_8(self):
        """
        Test that the worksheet is adjusted for 8 items.
        """
        item_count = 8
        setup_worksheet(self.worksheet, item_count)
        
        # Check if rows were inserted        
        # NOTE 
        # Row 23 is for the work order number
        # + 8 more rows for other information for the excel form 
        # = Row 31 as the last row containing some data in it
        self.assertEqual(self.worksheet.max_row, 31) 

        # Check if cells are merged correctly
        self.assertTrue(is_cell_merged(self.worksheet, "C27")) 
        self.assertTrue(is_cell_merged(self.worksheet, "B28"))
        self.assertTrue(is_cell_merged(self.worksheet, "J30"))

        # Check data validation
        dv_cells = [dv.sqref for dv in self.worksheet.data_validations.dataValidation]
        self.assertIn("B31", dv_cells)
        
    def test_setup_worksheet_more_than_8(self):
        """
        Test that the worksheet is adjusted for more than 8 items.
        """
        item_count = 10
        setup_worksheet(self.worksheet, item_count)
        
        # Check if rows were inserted        
        # NOTE 
        # Row 26 is for the work order number
        # + 8 more rows for other information for the excel form 
        # = Row 34 as the last row containing some data in it
        self.assertEqual(self.worksheet.max_row, 34) 

        # Check if cells are merged correctly
        self.assertFalse(is_cell_merged(self.worksheet, "C27"))
        self.assertFalse(is_cell_merged(self.worksheet, "B28"))
        self.assertTrue(is_cell_merged(self.worksheet, "J32"))

        # Check data validation
        dv_cells = [dv.sqref for dv in self.worksheet.data_validations.dataValidation]
        self.assertIn("B33", dv_cells)
        
    def tearDown(self):
        del self.workbook