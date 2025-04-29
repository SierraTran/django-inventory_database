"""
This module defines class-based views for displaying and managing items, item history, used items, 
item requests, and purchase order forms.

### Mixins
    - LoginRequiredMixin:
        Restricts access to authenticated users. Unauthenticated users will be redirected to the 
        login page. After logging in, they will be redirected back to the original destination 
        preserved by the query parameter defined by `redirect_field_name`.
    - SuperuserOrTechnicianRequiredMixin, SuperuserRequiredMixin, TechnicianRequiredMixin, 
    InternRequiredMixin, UserPassesTestMixin:
        Restricts access based on user-specific conditions.

### Base Classes
    - TemplateView:
        Renders a template with parameters from the URL included in the context.
    - ListView:
        Displays a list of objects.
    - DetailView:
        Shows the details of an object.
    - CreateView:
        Provides a form for creating new objects in the database.
    - UpdateView:
        Provides a form for updating existing objects in the database.
    - FormView:
        Displays a form and handles validation.
    - DeleteView:
        Confirms and processes object deletions.
"""

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import TemplateView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, FormView, DeleteView
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponse, HttpResponseForbidden, HttpResponseRedirect
from django.urls import reverse, reverse_lazy
from haystack.query import SearchQuerySet
from openpyxl import load_workbook
from inventory_database.mixins import (
    SuperuserOrTechnicianRequiredMixin,
    SuperuserRequiredMixin,
    TechnicianRequiredMixin,
    InternRequiredMixin,
)
from .forms import (
    ImportFileForm,
    PurchaseOrderItemFormSet,
    UsedItemForm,
    ItemRequestForm,
)
from .models import Item, ItemHistory, ItemRequest, PurchaseOrderItem, UsedItem
from .excel_functions import setup_worksheet


###################################################################################################
# Views for the Item Model ########################################################################
###################################################################################################
class ItemView(LoginRequiredMixin, ListView):
    """
    Class-based view to list all items.
    The user is required to be logged in to access this view.

    Inherits functionality from:
        - LoginRequiredMixin
        - ListView
    (See module docstring for more details on the inherited classes)

    Attributes:
        login_url (str): The URL to the login page (resolved using reverse_lazy).
        redirect_field_name (str): The query parameter for the URL the user will be redirected to 
            after logging in.
        model (Item): The model that this view will display.
        template_name (str): The name of the template to use for rendering the view.
        context_object_name (str): The name of the context variable to use for the list of items.

    Methods:
        `get_queryset()`: Retrieves the list of items to be displayed in alphanumeric order.
    """

    login_url = reverse_lazy("authentication:login")
    redirect_field_name = "next"
    model = Item
    template_name = "items.html"
    context_object_name = "items_list"

    def get_queryset(self):
        """
        Retrieves the list of items to be displayed in alphanumerical order by manufacturer, model,
        and part number.

        This method fetches all items from the database and orders them alphanumerically by 
        manufacturer, model, and part number.

        Returns:
            QuerySet: a queryset containing all items.
        """
        return Item.objects.all().order_by("manufacturer", "model", "part_number")


class ItemDetailView(LoginRequiredMixin, DetailView):
    """
    Class-based view for displaying the details of a single item.
    The user is required to be logged in to access this view.

    Inherits functionality from:
        - LoginRequiredMixin
        - DetailView
    (See module docstring for more details on the inherited classes)

    Attributes:
        login_url (str): The URL to the login page (resolved using reverse_lazy).
        redirect_field_name (str): The query parameter for the URL the user will be redirected to 
            after logging in.
        model (Item): The model that this view operates on.
        template_name (str): The template used to render the detail view.

    Methods:
        `get_context_data()`: Adds the user's group to the context data.
    """

    login_url = reverse_lazy("authentication:login")
    redirect_field_name = "next"
    model = Item
    template_name = "item_detail.html"

    def get_context_data(self, **kwargs):
        """
        Adds the user's group to the context data.

        This method calls the base class's `get_context_data` function to retrieve the base context
        data, obtains the first group the current user belongs to, and includes it in the context 
        data.

        Args:
            **kwargs: Additional keyword arguments passed to the parent method.

        Returns:
            dict: The context data, updated to include the user's group under the key "user_group".
        """
        context = super().get_context_data(**kwargs)
        context["user_group"] = self.request.user.groups.first()
        return context


class ItemCreateSuperuserView(SuperuserRequiredMixin, CreateView):
    """
    Class-based view for creating a new item.
    Only users in the "Superuser" group have access to this view.

    Inherits functionality from:
        - SuperuserRequiredMixin
        - CreateView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (Item): The model that this view operates on.
        fields (list): The fields to be displayed in the form.
        template_name (str): The name of the template used to render the view.

    Methods:
        `form_valid()`: Sets the `last_modified_by` field of the created Item as the current user.
    """

    model = Item
    fields = [
        "manufacturer",
        "model",
        "part_or_unit",
        "part_number",
        "description",
        "location",
        "quantity",
        "min_quantity",
        "unit_price",
    ]
    template_name = "item_create_form.html"

    def form_valid(self, form):
        """
        Overrides the form_valid function of the base class (`CreateView`) to pass the current user
        to the save method.

        This method sets the `last_modified_by` field of the new Item object to the current user 
        before calling the base class's `form_valid` method with the updated form.

        Args:
            form (ModelForm): The form that handles the data for updating the Item object.

        Returns:
            HttpResponse: The HTTP response object.
        """
        form.instance.last_modified_by = self.request.user
        return super().form_valid(form)


class ItemCreateTechnicianView(TechnicianRequiredMixin, CreateView):
    """
    Class-based view for creating a new item.
    This view requires the user to be in the 'Technician' group.

    Inherits functionality from:
        - TechnicianRequiredMixin
        - CreateView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (Item): The model that this view operates on.
        fields (list[str]): The fields to be displayed in the form.
        template_name (str): The name of the template used to render the view.

    Methods:
        `form_valid()`: Overrides form_valid to pass current user to save method
    """

    model = Item
    fields = [
        "manufacturer",
        "model",
        "part_or_unit",
        "part_number",
        "description",
        "location",
        "quantity",
        "unit_price",
    ]
    template_name = "item_create_form.html"

    def form_valid(self, form):
        """
        Overrides the form_valid function of the base class (`CreateView`) to pass the current user 
        to the save method.

        This method sets the `last_modified_by` field of the new Item object to the current user 
        before calling the base class's `form_valid` method with the updated form.

        Args:
            form (ModelForm): The form that handles the data for updating the Item object.

        Returns:
            HttpResponse: The HTTP response object.
        """
        form.instance.last_modified_by = self.request.user
        return super().form_valid(form)


class ItemUpdateSuperuserView(SuperuserRequiredMixin, UpdateView):
    """
    Class-based view for updating an existing item as a Superuser.
    This view requires the user to be in the "Superuser" group.

    Inherits functionality from:
        - SuperuserRequiredMixin
        - UpdateView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (Item): The model that this view operates on.
        fields (list[str]): The fields to be displayed in the form.
        template_name (str): The name of the template used to render the view.

    Methods:
        `form_valid()`: Passes the current user to the save method.
    """

    model = Item
    fields = [
        "manufacturer",
        "model",
        "part_or_unit",
        "part_number",
        "description",
        "location",
        "quantity",
        "min_quantity",
        "unit_price",
    ]
    template_name = "item_update_form.html"

    def form_valid(self, form):
        """
        Overrides the form_valid function of the base class `UpdateView` to pass the current user 
        to the save method.

        This method sets the `last_modified_by` field of the updated Item object to the current 
        user. Then, it calls the base class's `form_valid` method with the updated form.

        Args:
            form (ModelForm): The form that handles the data for updating the Item object.

        Returns:
            HttpResponse: The HTTP response object.
        """
        form.instance.save(user=self.request.user)
        return HttpResponseRedirect(self.get_success_url())


class ItemUpdateTechnicianView(TechnicianRequiredMixin, UpdateView):
    """
    Class-based view for updating an existing item as a Technician.
    This view requires the user to be in the "Technician" group.

    Inherits functionality from:
        - TechnicianRequiredMixin
        - UpdateView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (Item): The model that this view operates on.
        fields (list[str]): The fields to be displayed in the form.
        template_name (str): The name of the template used to render the view.

    Methods:
        `form_valid()`: Overrides form_valid function of the `UpdateView` base class to pass the 
            current user to the save method.
    """

    model = Item
    fields = [
        "manufacturer",
        "model",
        "part_or_unit",
        "part_number",
        "description",
        "location",
        "quantity",
        "unit_price",
    ]
    template_name = "item_update_form.html"

    def form_valid(self, form):
        """
        Overrides form_valid function of the `UpdateView` base class to pass the current user to 
        the save method.

        This method sets the `last_modified_by` field of the updated Item object to the current 
        user. Then, it calls the base class's `form_valid` method with the updated form.

        Args:
            form (ModelForm): The form that handles the data for updating the Item object.

        Returns:
            HttpResponse: The HTTP response object that's returned to the client.
        """
        form.instance.updated_by = self.request.user
        return super().form_valid(form)


class ItemUpdateInternView(InternRequiredMixin, UpdateView):
    """
    Class-based view for updating the quantity of an existing item as an Intern.
    This view requires the user to be in the "Intern" group.

    Inherits functionality from:
        - InternRequiredMixin
        - UpdateView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (Item): The model that this view operates on.
        fields (list[str]): The fields to be displayed in the form. For interns, only the quantity 
            is available to them.
        template_name (str): The name of the template used to render the view.

    Methods:
        `form_valid()`: Overrides the `form_valid` function of the base class `UpdateView` to pass 
            the current user to the save method.
    """

    model = Item
    fields = ["quantity"]
    template_name = "item_update_form.html"

    def form_valid(self, form):
        """
        Overrides the form_valid function of the base class `UpdateView` to pass the current user 
        to the save method.

        This method sets the `last_modified_by` field of the updated Item object to the current 
        user. Then, it calls the base class's `form_valid` method with the updated form.

        Args:
            form (ModelForm): The form that handles the data for updating the Item object.

        Returns:
            HttpResponse: The HTTP response object.
        """
        form.instance.last_modified_by = self.request.user
        return super().form_valid(form)


class ItemDeleteView(SuperuserOrTechnicianRequiredMixin, DeleteView):
    """
    Class-based view for deleting an existing item.
    This view requires the user to be in the "Superuser" or "Technician" group.

    Inherits functionality from:
        - SuperuserOrTechnicianRequiredMixin
        - DeleteView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (Item): The model that this view operates on.
        template_name (str): The name of the template that the view will render.
        success_url (str): The URL to redirect to upon successful deletion.
        fail_url (str): The URL to redirect to if the deletion is canceled.

    Methods:
        `get_fail_url()`: Returns the URL to redirect to if the deletion is canceled.
        `post()`: Handles POST requests to delete the item or cancel the deletion.
    """

    model = Item
    template_name = "inventory/item_confirm_delete.html"
    success_url = reverse_lazy("inventory:items")

    def get_fail_url(self):
        """
        Returns the URL to redirect to if the deletion is canceled.

        This method uses reverse_lazy to resolve the failure URL with the primary key (pk) of the 
        object being processed and returns it.

        Returns:
            str: The URL to redirect to.
        """
        return reverse_lazy(
            "inventory:item_detail", kwargs={"pk": self.get_object().pk}
        )

    fail_url = property(get_fail_url)

    def post(self, request, *args, **kwargs):
        """
        Handles POST requests to delete the item or cancel the deletion.

        This method first checks which button was pressed in the form. If the "Cancel" button was 
        pressed, the user is redirected back to the Item Detail page (the failure URL). If the 
        "Confirm" button was pressed (the else case), the item is deleted. If there are other 
        objects that reference the item, an IntegrityError is caught and an error message is 
        displayed after the user is redirected to the item Detail page.

        Args:
            request (HttpRequest): The HTTP request object containing metadata about the request.
            *args: Additional positional arguments.
            **kwargs: Additional keyword arguments.

        Returns:
            HttpResponse: The HTTP response object.
        """
        if "Cancel" in request.POST:
            return redirect(self.fail_url)
        return super(ItemDeleteView, self).post(request, *args, **kwargs)


class SearchItemsView(LoginRequiredMixin, ListView):
    """
    Class-based view for searching items.
    This view uses Haystack to perform the search.

    Inherits functionality from:
        - LoginRequiredMixin
        - ListView
    (See module docstring for more details on the inherited classes)

    Attributes:
        login_url (str): The URL to the login page (resolved using reverse_lazy).
        redirect_field_name (str): The query parameter for the URL the user will be redirected to 
            after logging in.
        model (Item): The model that this view operates on.
        template_name (str): The template used to render the search results.
        context_object_name (str): The name of the context variable to use for the search results.

    Methods:
        get_queryset(): Retrieves the search results based on the query.
    """

    login_url = reverse_lazy("authentication:login")
    redirect_field_name = "next"
    model = Item
    template_name = "search/item_search.html"
    context_object_name = "results_list"

    def get_queryset(self):
        """
        Retrieves search results based on the query parameter.

        This method extracts the search query from the GET request (`q` parameter), filters the
        search queryset for objects containing the query term, and sorts the results by
        "manufacturer", "model", and "part_number".

        Returns:
            list: A list of filtered and ordered search results. Returns an empty list if no query 
                is provided.
        """
        query = self.request.GET.get("q")
        results = (
            SearchQuerySet()
            .filter(content=query)
            .order_by("manufacturer", "model", "part_number")
            if query
            else []
        )
        return results


class ImportItemDataView(SuperuserOrTechnicianRequiredMixin, FormView):
    """
    Renders a view to allow users to import items from an .xlsx file to the database.
    Users must be in the "Superuser" or Technician" group to access this view.

    Inherits functionality from:
        - SuperuserOrTechnicianRequiredMixin
        - FormView
    (See module docstring for more details on the inherited classes)

    Attributes:
        form_class (Form): The form that the view operates on.
        template_name (str): The name of the template to be rendered by the class.
        success_url (str): The URL to redirect to after the form is successfully processed.

    Methods:
        `form_valid(form)`: Processes data from an uploaded Excel file to the database.
    """

    form_class = ImportFileForm
    template_name = "import_item_data.html"
    success_url = reverse_lazy("inventory:items")

    def form_valid(self, form) -> HttpResponseRedirect:
        """
        Processes the uploaded Excel file from the form, extracts item data from each row,
        and creates Item objects in the database. Empty cells will have a default value set
        for them in the database.

        Args:
            form (Form): The form containing the uploaded Excel file.

        Returns:
            HttpResponseRedirect: The HTTP response to redirect to the items list view after 
                processing the file.
        """
        file = form.cleaned_data["file"]
        workbook = load_workbook(file)
        sheet = workbook.active
        user = self.request.user

        # For each record in the excel file ...
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # If the row is completely blank, stop the for loop
            if all(cell is None for cell in row):
                break

            # If not...
            # Get its data. Set to default value if None
            manufacturer = row[0] if row[0] is not None else "N/A"
            model = row[1] if row[1] is not None else "N/A"
            part_or_unit = row[2] if row[2] is not None else "Part"
            part_number = row[3] if row[3] is not None else ""
            description = row[4] if row[4] is not None else ""
            location = row[5] if row[5] is not None else "N/A"
            quantity = row[6] if row[6] is not None else 0
            min_quantity = row[7] if row[7] is not None else 0
            unit_price = row[8] if row[8] is not None else 0.01

            # Create a new Item with the data
            Item.objects.create(
                manufacturer=manufacturer,
                model=model,
                part_or_unit=part_or_unit,
                part_number=part_number,
                description=description,
                location=location,
                quantity=quantity,
                min_quantity=min_quantity,
                unit_price=unit_price,
                last_modified_by=user,
            )

        # Go to items page after finishing
        return HttpResponseRedirect(reverse("inventory:items"))


###################################################################################################
# Views for the ItemHistory Model #################################################################
###################################################################################################
class ItemHistoryView(LoginRequiredMixin, ListView):
    """
    Class-based view for displaying the history of a specific item.
    This view requires the user to be logged in.

    Inherits functionality from:
        - LoginRequiredMixin
        - ListView
    (See module docstring for more details on the inherited classes)

    Attributes:
        login_url (str): The URL to the login page (resolved using reverse_lazy).
        redirect_field_name (str): The query parameter for the URL the user will be redirected to 
            after logging in.
        model (ItemHistory): The model that this view operates on.
        template_name (str): The template used to render the history view.
        context_object_name (str): The context variable name for the list of item history records.

    Methods:
        `get_queryset()`: Retrieves the history records for the specific item in reverse 
            chronological order.
        `get_context_data()`: Adds the specific item to the context data.
    """

    login_url = reverse_lazy("authentication:login")
    redirect_field_name = "next"
    model = ItemHistory
    template_name = "item_history.html"
    context_object_name = "item_history_list"

    def get_queryset(self):
        """
        Retrieves the history records for the specific item in reverse chronological order.

        This method extracts the ID of the item from the URL parameters, filters the `ItemHistory` 
        objects to match the given item ID, and orders the resulting queryset by the `timestamp` 
        field in descending order (most recent first).

        Returns:
            QuerySet: A queryset containing the history records for the specified item in reverse 
                chronological order.
        """
        item_id = self.kwargs["pk"]
        history = ItemHistory.objects.filter(item_id=item_id).order_by("-timestamp")
        return history

    def get_context_data(self, **kwargs):
        """
        Adds the specific item to the context data.

        This method extracts the ID of the item from the URL parameters, calls the base class's
        `get_context_data` method to get the base context, and then fetches the specific Item
        object with the matching ID before including it in the context data.

        If no Item object is found with the given ID, an `Http404` exception is raised.

        Args:
            **kwargs: Additional keyword arguments passed to the parent method.

        Returns:
            dict: The context data with the specific item added.
        """
        item_id = self.kwargs["pk"]
        context = super().get_context_data(**kwargs)
        context["item"] = get_object_or_404(Item, id=item_id)
        return context


###################################################################################################
# Views for the ItemRequest Model #################################################################
###################################################################################################
class ItemRequestView(SuperuserOrTechnicianRequiredMixin, ListView):
    """
    Class-based view for displaying item requests.
    Only users belonging to the "Technician" or "Superuser" groups are allowed to access this view.

    Inherits functionality from:
        - SuperuserOrTechnicianRequiredMixin
        - ListView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (ItemRequest): The model that this view will display.
        template_name (str): The template used to render the view.
        context_object_name (str): The context variable name for the list of item requests.

    Methods:
        `get_queryset()`: Returns the queryset of all item requests.
    """

    model = ItemRequest
    template_name = "item_requests.html"
    context_object_name = "item_requests_list"

    def get_queryset(self):
        """
        Retrieves all Item Requests from the database.

        Returns:
            QuerySet: The queryset containing all item requests.
        """
        return ItemRequest.objects.all().order_by("timestamp")


class ItemRequestDetailView(SuperuserOrTechnicianRequiredMixin, DetailView):
    """
    Class-based view for displaying the details of a ItemRequest.
    Users must be in the "Technician" or "Superuser" group to access this view.

    Inherits functionality from:
        - SuperuserOrTechnicianRequiredMixin
        - DetailView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (ItemRequest): The model that the view will operate on.
        template_name (str): The template that will be used to render the view.

    Methods:
        `get_context_data()`:  Adds the name of the current user's group to the context.
    """

    model = ItemRequest
    template_name = "item_request_detail.html"

    def get_context_data(self, **kwargs):
        """
        Adds the name of the current user's group to the context.

        This method retrieves the base context by calling the base class's `get_context_data` 
        method. Then, it retrieves the name of the first group the current user belongs to 
        and adds it to the context data under the "current_user_group_name" key.

        Arguments:
            **kwargs: Additional keyword arguments.

        Returns:
            dict: The context data, including the name of the current user's first group, for use 
                in the view.
        """
        context = super().get_context_data(**kwargs)
        context["current_user_group_name"] = self.request.user.groups.first().name
        return context


class ItemRequestCreateView(TechnicianRequiredMixin, CreateView):
    """
    Class-based view for creating an item request.
    This view requires the user to be in the "Technician" group.

    Inherits functionality from:
        - TechnicianRequiredMixin
        - CreateView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (ItemRequest): The model that this view operates on.
        fields (list): The fields to be displayed in the form.
        template_name (str): The name of the template used to render the view.

    Methods:
        `get_initial()`: Retrieves initial item data from the GET parameters and the request.
        `get_context_data()`: Adds the specific item to the context data.
    """

    model = ItemRequest
    form_class = ItemRequestForm
    template_name = "item_request_form.html"

    def get_initial(self):
        """
        Retrieves initial item data from the GET parameters and the request.

        This method calls the base class's `get_initial` method to get the base initial data. Then,
        it extracts the manufacturer, model_part_num, description, and unit_price from the GET 
        parameters under the keys "manufacturer", "model_part_num", "description", and 
        "unit_price", respectively. After that, the current user is saved under the "requested_by" 
        key. The initial data is then returned.

        Returns:
            dict: The initial data for the form.
        """
        initial = super().get_initial()
        initial["manufacturer"] = self.request.GET.get("manufacturer", "")
        initial["model_part_num"] = self.request.GET.get("model_part_num", "")
        initial["description"] = self.request.GET.get("description", "")
        initial["unit_price"] = self.request.GET.get("unit_price", "")
        return initial

    def get_context_data(self, **kwargs):
        """
        Adds the specific item to the context data.

        This method retrieves the base context by calling the base class's `get_context_data` 
        method. Then, it obtains the "item_id" through the GET parameters of the request. Finally, 
        it fetches the `Item` object with the provided ID and adds it to the context under the 
        "item" key. If no `Item` object is found, an `Http404` exception is raised. The context 
        data is then returned.

        Args:
            **kwargs: Additional keyword arguments ot pass to the base class.

        Returns:
            dict: The context data with the specific item added.
        """
        context = super().get_context_data(**kwargs)
        item_id = self.request.GET.get("item_id")
        if item_id:
            context["item"] = get_object_or_404(Item, pk=item_id)
        else:
            context["item"] = None
        return context

    def form_valid(self, form):
        """
        Overrides the form_valid function of the base class (`CreateView`) to pass the current user
        to the save method.

        This method sets the `requested_by` field of the new ItemRequest object to the current user
        before calling the base class's `form_valid` method with the updated form.

        Args:
            form (ModelForm): The form that handles the data for creating the ItemRequest object.

        Returns:
            HttpResponse: The HTTP response object.
        """
        form.instance.requested_by = self.request.user
        return super().form_valid(form)


class ItemRequestAcceptView(SuperuserRequiredMixin, TemplateView):
    """
    Class-based view for confirming or canceling the acceptance of an item request.
    Only users in the "Superuser" group can access this view.

    Inherits functionality from:
        - SuperuserRequiredMixin
        - TemplateView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (ItemRequest): The model that this view operates on.
        template_name (str): The name of the template used to render the view.
        fail_url (str): The URL to redirect to if the acceptance is canceled.

    Methods:
        `get_object()`: Retrieves the specific ItemRequest object for the view.
        `get_fail_url()`: Returns the URL to redirect to if the acceptance is canceled.
        `get_context_data()`: Adds the specific item request to the context data.
        `post()`: Handles POST requests to set the item request's status to "Accepted" or cancel
            the operation.
    """

    model = ItemRequest
    template_name = "item_request_confirm_accept.html"

    def get_object(self):
        """
        Retrieves the specific ItemRequest object for the view.

        This method fetches the ItemRequest object with the primary key (pk) extracted from the
        `kwargs` using the `get_object_or_404` function. If no ItemRequest object is found with
        the given primary key, an `Http404` exception is raised.

        Returns:
            ItemRequest: The ItemRequest object that may or may not be accepted by a Superuser.
        """
        return get_object_or_404(ItemRequest, pk=self.kwargs.get("pk"))

    def get_fail_url(self):
        """
        Resolves the URL to redirect to if the acceptance is canceled.

        This method uses reverse_lazy to resolve the failure URL with the primary key (pk) of the
        object being processed and returns it.

        Returns:
            str: The resolved URL for redirction.
        """
        return reverse_lazy(
            "inventory:item_request_detail", kwargs={"pk": self.get_object().pk}
        )

    fail_url = property(get_fail_url)

    def get_context_data(self, **kwargs):
        """
        Adds the specific item request to the context data.

        This method retrieves the base context by calling the base class's `get_context_data`
        method. Then, it adds the object returned by `get_object` method to the context under 
        the "object" key.

        Returns:
            dict: The context data for the view including the specific item request.
        """
        context = super().get_context_data(**kwargs)
        context["object"] = self.get_object()
        return context

    def post(self, request, *args, **kwargs):
        # NOTE: Although this function doesn't use *args or **kwargs,
        # they need to be included to avoid errors.
        """
        Handles POST requests to set the item request's status to "Accepted" or cancel the
        operation.

        This method checks the submitted form data to determine if the operation should be 
        canceled (redirecting to the failure URL) or if the item request's status should be
        updated to "Accepted". If the item request's status is updated, the object will be
        saved and the user will be redirected to the item request's detail page.

        Arguments:
            request (HttpRequest): The HTTP request object containing POST data.

        Returns:
            HttpResponseRedirect: A redirect response after canceling or confirming the status 
                change.
        """
        if "Cancel" in request.POST:
            return redirect(self.fail_url)

        item_request = self.get_object()
        item_request.status = "Accepted"
        item_request.status_changed_by = self.request.user
        item_request.save()
        return redirect(item_request.get_absolute_url())


class ItemRequestRejectView(SuperuserRequiredMixin, TemplateView):
    """
    Class-based view for confirming or canceling the acceptance of an item request.
    Only users in the 'Superuser' group can access this view.

    Inherits functionality from:
        - SuperuserRequiredMixin
        - TemplateView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (ItemRequest): The model that this view operates on.
        template_name (str): The name of the template used to render the view.
        fail_url (str): The URL to redirect to if the rejection is canceled.

    Methods:
        `get_object()`: Retrieves the specific item request for the view.
        `get_fail_url()`: Returns the URL to redirect if the rejection is canceled.
        `get_context_data()`: Adds the specific item request to the context data.
        `post()`: Handles POST requests to set the item request's status to "Rejected" or cancel the 
            operation.
    """

    model = ItemRequest
    template_name = "item_request_confirm_reject.html"

    def get_object(self):
        """
        Retrieves the specific item request for the view.

        This method retrieves the primary key (pk) from `kwargs` and then fetches the `ItemRequest` 
        object with the matching primary key. If no `ItemRequest` object is found with the given 
        primary key, an `Http404` exception is raised.

        Returns:
            ItemRequest: The item request that may or may not be rejected by a Superuser.
        """
        return get_object_or_404(ItemRequest, pk=self.kwargs.get("pk"))

    def get_fail_url(self):
        """
        Resolves the URL to redirect to if the rejection is canceled.

        This method uses reverse_lazy to resolve the failure URL with the primary key (pk) of the 
        object being processed and returns it.

        Returns:
            str: The resolvd URL for redirection.
        """
        return reverse_lazy(
            "inventory:item_request_detail", kwargs={"pk": self.get_object().pk}
        )

    fail_url = property(get_fail_url)

    def get_context_data(self, **kwargs):
        """
        Adds the specific item request to the context data.

        This method retrieves the base context by calling the base class's `get_context_data` 
        method. Then, it adds the object returned by `get_object` method to the context under 
        the "object" key.

        Args:
            **kwargs: Additional keyword arguments passed to the base class.

        Returns:
            dict: The context data for the view including the specific item request.
        """
        context = super().get_context_data(**kwargs)
        context["object"] = self.get_object()
        return context

    def post(self, request, *args, **kwargs):
        # NOTE: Although this function doesn't use *args or **kwargs, they need to be included to
        # avoid errors.
        """
        Handles POST requests to set the item request's status to "Rejected" or cancel the 
        operation.

        This method checks the submitted form data to determine if the operation should be canceled
        (redirecting to the failure URL) or if the item request's status should be updated to 
        "Rejected". If the item request's status is updated, the object will be saved and the user 
        will be redirected to the item request's detail page.

        Arguments:
            request (HttpRequest): The HTTP request object containing POST data.

        Returns:
            HttpResponseRedirect: A redirect response after canceling or confirming the status 
                change.
        """
        if "Cancel" in request.POST:
            return redirect(self.fail_url)

        item_request = self.get_object()
        item_request.status = "Rejected"
        item_request.save()
        return redirect(item_request.get_absolute_url())


class ItemRequestDeleteView(UserPassesTestMixin, DeleteView):
    """
    Class-based view for confirming or canceling the deletion of an item request.
    Only users who made the request can access this view.

    Inherits functionality from:
        - UserPassesTestMixin
        - DeleteView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (ItemRequest): The model that this view operates on.
        template_name (str): The name of the template used to render the view.
        success_url (str): The URL to redirect to upon successful deletion.
        fail_url (str): The URL to redirect to if the deletion is canceled.

    Methods:
        `test_func()`: Checks if the item request belongs to the user.
        `handle_no_permission()`: Renders the 403 page with a message explaining the reason for the 
            error.
        `get_fail_url()`: Returns the URL to redirect if the deletion is canceled.
        `post()`: Handles POST requests to delete the item or cancel the deletion.
    """

    model = ItemRequest
    template_name = "item_request_confirm_delete.html"
    success_url = reverse_lazy("inventory:item_requests")

    def test_func(self):
        """
        Checks if the item request belongs to the user.

        This method retrieves the primary key (pk) from `kwargs` and then fetches the `ItemRequest`
        object with the matching primary key. If no object is found, an `Http404` exception is 
        raised. Then, it checks if the `requested_by` field of the `ItemRequest` object matches the
        current user. If it does, True is returned, indicating that the user is allowed to delete 
        the item request. Otherwise, False is returned.

        Returns:
            bool: True if the user is the one who made the item request. False otherwise.
        """
        user = self.request.user
        item_request_id = self.kwargs.get("pk")
        request_from = get_object_or_404(ItemRequest, id=item_request_id).requested_by
        return request_from == user

    def handle_no_permission(self):
        """
        Renders the 403 page with a message explaining the reason for the error.

        Returns:
            HttpResponse: The HTTP response object with the rendered 403 page.
        """
        message = "You didn't make this item request, so you can't delete it. Please ask the author of the item request to delete it."
        return HttpResponseForbidden(
            render(self.request, "403.html", {"message": message})
        )

    def get_fail_url(self):
        """
        Returns the URL to redirect to if the deletion is canceled.

        This method uses reverse_lazy to resolve the failure URL with the primary key (pk) of the 
        object being processed and returns it.

        Returns:
            str: The URL to redirect to.
        """
        return reverse_lazy("inventory:item_requests")

    fail_url = property(get_fail_url)

    def post(self, request, *args, **kwargs):
        """
        Handles POST requests to delete the item or cancel the deletion.

        This method first checks which button was pressed in the form. If the "Cancel" button was 
        pressed, the user is redirected back to the Item Requests page (the failure URL). If the 
        "Confirm" button was pressed (the else case), the item request is deleted.

        Args:
            request (HttpRequest): The HTTP request object containing metadata about the request.
            *args: Additional positional arguments.
            **kwargs: Additional keyword arguments.

        Returns:
            HttpResponse: The HTTP response object.
        """
        if "Cancel" in request.POST:
            return redirect(self.fail_url)
        return super(ItemRequestDeleteView, self).post(request, *args, **kwargs)


###################################################################################################
# Views for the UsedItem Model ####################################################################
###################################################################################################
class UsedItemView(LoginRequiredMixin, ListView):
    """
    Class-based view for displaying all Used Items.
    Users must be logged in to have access to this view.

    Inherits functionality from:
        - LoginrequiredMixin
        - ListView
    (See module docstring for more details on the inherited classes)

    Attributes:
        login_url (str): The URL to the login page (resolved using reverse_lazy).
        redirect_field_name (str): The query parameter for the URL the user will be redirected to
            after logging in.
        model (UsedItem): The model that the view will operate on.
        template_name (str): The template that will be used to render the view.
        context_object_name (str): The name of the context object.

    Methods:
        `get_queryset()`: Retrieves all UsedItems from the database in order of their work order 
            and item.
    """

    login_url = reverse_lazy("authentication:login")
    redirect_field_name = "next"
    model = UsedItem
    template_name = "used_items.html"
    context_object_name = "used_items_list"

    def get_queryset(self):
        """
        Retrieves all UsedItems from the database in order of their work order and item.

        This method retrieves all UsedItem objects from the database and orders them by work_order
        and item.

        Returns:
            QuerySet: A queryset containing all used items.
        """
        return UsedItem.objects.all().order_by("-datetime_used", "work_order", "item")


class UsedItemDetailView(LoginRequiredMixin, DetailView):
    """
    Class-based view to display the details for a specific used item.
    Users must be logged in to have access to this view.

    Inherits functionality from:
        - LoginrequiredMixin
        - DetailView
    (See module docstring for more details on the inherited classes)

    Attributes:
        login_url (str): The URL to the login page (resolved using reverse_lazy).
        redirect_field_name (str): The query parameter for the URL the user will be redirected to 
            after logging in.
        model (UsedItem): The model on which the view will operate.
        template_name (str): The template that will be used to render the view.

    Methods:
        `get_queryset()`: Adds the specific used item to the context data.
    """

    login_url = reverse_lazy("authentication:login")
    redirect_field_name = "next"
    model = UsedItem
    template_name = "used_item_detail.html"

    def get_context_data(self, **kwargs):
        """
        Adds the specific used item to the context data.

        This method retrieves the base context data by calling the base class's `get_context_data`
        function. Then, it adds the specific `UsedItem` object, represented by `self.object`, to 
        the context under the "used_item" key.

        Args:
            **kwargs: Additional keyword arguments passed to the base class's `get_context_data` 
                method.

        Returns:
            dict: The context data, including the used item, for use in the template.
        """
        context = super().get_context_data(**kwargs)
        context["used_item"] = self.object
        return context


class UsedItemCreateView(SuperuserOrTechnicianRequiredMixin, CreateView):
    """
    Class-based view for displaying the page to create a Used Item.
    Only users in the "Superuser" and "Technician" group have access to this view.

    Inherits functionality from:
        - SuperuserOrTechnicianRequiredMixin
        - CreateView
    (See module docstring for more details on the inherited classes)

    Attributes:
        model (UsedItem): The model on which the view will operate.
        fields (str): The fields to be displayed in the view.
        template_name (str): The template that will be used to render the view.

    Methods:
        `get_initial()`: Adds the specific item to the initial data to be used in the form.
        `get_context_data()`: Adds the item and initial data for the form to the context.
        `dispatch()`: Checks if the item's quantity is greater than 0 before allowing access to the
            view.
        `form_valid()`: Decrements the quantity of the associated Item when a new UsedItem is 
            created and updates the ItemHistory record to explain the decrement.
    """

    model = UsedItem
    form_class = UsedItemForm
    template_name = "item_use_form.html"

    def get_initial(self):
        """
        Adds the specific item and current user to the initial data to be used in the form.

        This method retrieves the base initial data by calling the base class's `get_initial`
        function. Then, it adds the current user to the initial data under they "used_by" key. If 
        an "item_id" is detected in the GET parameters, the corresponding `Item` object is 
        retrieved and added to the initial data under the "item" key.

        Returns:
            dict: The initial data for creating a Used Item, including the user and, if applicable,
                the specified item.
        """
        initial = super().get_initial()
        current_user = self.request.user
        initial.update({"used_by": current_user})
        item_id = self.request.GET.get("item_id")
        if item_id:
            item = get_object_or_404(Item, pk=item_id)
            initial.update(
                {
                    "item": item,
                }
            )
        return initial

    def get_context_data(self, **kwargs):
        """
        Adds the item and initial data for the form to the context.

        This method retrieves the base context data by calling the base class's `get_context_data`
        function. Then, if an "item_id" is detected in the GET parameters, the corresponding Item 
        object is retrieved and added to the context data under the "item" key. If the request 
        method is GET, the form in the context (under the "form" key) has itsinitial data updated 
        with values from `get_initial`.

        Args:
            **kwargs: Additional keyword arguments passed to the base class's method.

        Returns:
            dict: The context data including the specific item an updated initial data for the form.
        """
        context = super().get_context_data(**kwargs)
        item_id = self.request.GET.get("item_id")
        item = get_object_or_404(Item, pk=item_id)
        context["item"] = item
        if self.request.method == "GET":
            form = context["form"]
            form.initial.update(self.get_initial())
        return context

    def dispatch(self, request, *args, **kwargs):
        """
        Checks if the item's quantity is greater than 0 before allowing access to the view.

        This method first retrieves the item_id from the GET parameters and then retrieves the Item
        object with the corresponding ID. If the quantity of the item is less than or equal to 0, 
        an error message is displayed and the user is redirected to the detail page for the item.
        Otherwise, the request is dispatched to theof base class's `dispatch` method.

        Args:
            request (HttpRequest): The HTTP request object.
            *args: Additional positional arguments.
            **kwargs: Additional keyword arguments.

        Returns:
            HttpResponse: The HTTP response object.
        """
        item_id = self.request.GET.get("item_id")
        item = get_object_or_404(Item, pk=item_id)
        if item.quantity <= 0:
            messages.error(request, "Cannot use item with quantity 0.")
            return redirect("inventory:item_detail", pk=item_id)
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        """
        Decrements the quantity of the associated Item when a new UsedItem is created and updates
        the ItemHistory record to explain the decrement.

        This method first calls the base class's `form_valid` method to process the form data. 
        Then, it retrieves the UsedItem from the form and decremented the quantity of the
        associated Item by 1 and saves it. Next, the URL of the UsedItem is resolved with `reverse`
        with its primary key (pk) and stored in `used_item_url`. After that, the last ItemHistory 
        record, which was created when the Item was saved, is retrieved and updated with the
        action "use" and a link to where the UsedItem is used in the changes field. Finally, it
        saves the history record and returns the response.

        Args:
            form (ModelForm): The form that handles the data for creating a new UsedItem object.

        Returns:
            HttpResponse: The HTTP response object.
        """
        response = super().form_valid(form)
        used_item = form.instance
        item = used_item.item
        item.quantity -= 1
        item.last_modified_by = self.request.user
        item.save()

        used_item_url = reverse(
            "inventory:used_item_detail", kwargs={"pk": used_item.pk}
        )

        history_record_to_edit = ItemHistory.objects.last()
        history_record_to_edit.action = "use"
        history_record_to_edit.changes += f', <a href="{used_item_url}">Item used in work order {used_item.work_order}</a>'
        history_record_to_edit.save()

        return response


class UsedItemDeleteView(SuperuserOrTechnicianRequiredMixin, DeleteView):
    """
    Class-based view for confirming or canceling the deletion of a used item.
    Only users in the "Superuser" or "Technician" group can access this view.

    Inherits functionality from:
        - SuperuserOrTechnicianRequiredMixin
        - DeleteView

    Attributes:
        model (UsedItem): The model that this view operates on.
        template_name (str): The name of the template used to render the view.
        success_url (str): Redirection URL if deletionis confirmed
        fail_url (str): Redirection URL if deletion is canceled

    Methods:
        `get_fail_url()`: Returns the URL to redirect if the deletion is canceled.
        `post()`: Handles POST requests to delete the used item or cancel the deletion.
    """
    model = UsedItem
    template_name = "used_item_confirm_delete.html"
    success_url = reverse_lazy("inventory:used_items")

    def get_fail_url(self):
        """
        Returns the URL to redirect to if the deletion is canceled.

        This method uses reverse_lazy to resolve the failure URL with the primary key (pk) of the
        object being processed and returns it.

        Returns:
            str: The URL to redirect to.
        """
        return reverse_lazy("inventory:used_items")

    fail_url = property(get_fail_url)

    def post(self, request, *args, **kwargs):
        """
        Handles POST requests to delete the item or cancel the deletion.

        This method first checks which button was pressed in the form. If the "Cancel" button was
        pressed, the user is redirected back to the Used Items page (the failure URL). If the 
        "Confirm" button was pressed (the else case), the used item is deleted.

        Args:
            request (HttpRequest): The HTTP request object containing metadata about the request.
            *args: Additional positional arguments.
            **kwargs: Additional keyword arguments.

        Returns:
            HttpResponse: The HTTP response object.
        """
        if "Cancel" in request.POST:
            return redirect(self.fail_url)
        return super(UsedItemDeleteView, self).post(request, *args, **kwargs)


class SearchUsedItemsView(LoginRequiredMixin, ListView):
    """
    Class-based view for searching used items.
    This view uses Haystack to perform the search.

    Inherits functionality from:
        - LoginRequiredMixin
        - ListView
    (See module docstring for more details on the inherited classes)

    Attributes:
        login_url (str): The URL to the login page (resolved using reverse_lazy).
        redirect_field_name (str): The query parameter for the URL the user will be redirected to 
            after logging in.
        model (UsedItem): The model that this view operates on.
        template_name (str): The template used to render the search results.
        context_object_name (str): The name of the context variable to use for the search results.

    Methods:
        `get_queryset()`: Retrieves the search results based on the query.
    """

    login_url = reverse_lazy("authentication:login")
    redirect_field_name = "next"
    model = UsedItem
    template_name = "search/used_item_search.html"
    context_object_name = "results_list"

    def get_queryset(self):
        """
        Retrieves the search results based on the query.

        This method extracts the search query from the GET request (`q` parameter), filters the
        search queryset for objects containing the query term, and sorts the results by
        `work_order` and `item`.

        Returns:
            list: A list of search results.
        """
        query = self.request.GET.get("q")
        results = (
            SearchQuerySet()
            .models(UsedItem)
            .filter(content=query)
            .order_by("work_order", "item")
            if query
            else []
        )
        return results


###################################################################################################
# Views for the PurchaseOrderItem model ###########################################################
###################################################################################################
class PurchaseOrderItemsFormView(SuperuserRequiredMixin, FormView):
    """
    Renders a view to allow users to create purchase orders using a formset.
    Users must be in the "Superuser" group to access this view.

    Inherits functionality from:
        - SuperuserRequiredMixin
        - FormView
    (See module docstring for more details on the inherited classes)

    Attributes:
        form_class (FormSet): The formset class to use for the purchase order items.
        template_name (str): The template used to render the formset.
        success_url (str): The URL to redirect to upon successful form submission.

    Methods:
        `get_context_data()`: Adds the formset to the context data.
        `form_valid()`: Processes the formset data and writes it to an Excel file for download.
    """

    form_class = PurchaseOrderItemFormSet
    template_name = "purchase_order_form.html"
    success_url = reverse_lazy("inventory:items")

    def get_initial(self):
        """
        Returns the initial data to use for the formset.

        This method retrieves initial data from the GET parameters and returns it as a list of 
        dictionaries, each representing the initial data for one form in the formset.

        Returns:
            list: A list of dictionaries containing the initial data for the formset.
        """
        initial_data = []
        manufacturer = self.request.GET.get("manufacturer", "")
        model_part_num = self.request.GET.get("model_part_num", "")
        quantity_ordered = self.request.GET.get("quantity_ordered", 1)
        description = self.request.GET.get("description", "")
        unit_price = self.request.GET.get("unit_price", 0.00)

        initial_data.append(
            {
                "manufacturer": manufacturer,
                "model_part_num": model_part_num,
                "quantity_ordered": quantity_ordered,
                "description": description,
                "unit_price": unit_price,
            }
        )

        return initial_data

    def get_context_data(self, **kwargs):
        """
        Adds the formset and query parameters to the context data.

        This method first retrieves the base context data by calling the base class's 
        `get_context_data` method. Then, if a POST request is detected, the submitted form data is
        added as a formset to the context under the "formset" key. Otherwise, an empty queryset is
        used to initialize the `PurchaseOrderItemFormSet`, which is also added to the context under
        the "formset" key.

        Args:
            **kwargs: Additional keyword arguments passed to the base class's `get_context_data` 
                method.

        Returns:
            dict: The context data, including the formset and query parameters.
        """
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context["formset"] = PurchaseOrderItemFormSet(self.request.POST)
        else:
            context["formset"] = PurchaseOrderItemFormSet(
                initial=self.get_initial(),
                queryset=PurchaseOrderItem.objects.none(),
            )
        return context

    def form_valid(self, formset):
        """
        Processes the formset data and writes it to an Excel file for download.

        This method is called when valid form data has been POSTed. It writes the purchase order 
        data from the formset to an Excel file using a predefined template and returns an HTTP 
        response to download the generated Excel file.

        Args:
            formset (FormSet): The formset containing the purchase order data.

        Returns:
            HttpResponse: The HTTP response object to download the Excel file.
        """
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            'attachment; filename="new_purchase_order.xlsx"'
        )
        po_template_path = "PO_Template.xlsx"
        workbook = load_workbook(po_template_path)
        worksheet = workbook.active

        # Count the items in the form
        item_count = 0
        for form in formset:
            # Skip forms marked for deletion
            if form.cleaned_data.get("DELETE"):
                continue
            item_count += 1

        # If there are more than 8 items, set up the worksheet to accommodate them
        if item_count > 8:
            setup_worksheet(worksheet, item_count)

        # Write data to the worksheet
        # In the worksheet, the first item row is 16
        row = 16
        for form in formset:
            # Ignore forms that are marked for deletion
            if form.cleaned_data.get("DELETE"):
                continue

            manufacturer = form.cleaned_data["manufacturer"]
            model_part_num = form.cleaned_data["model_part_num"]
            quantity_ordered = form.cleaned_data["quantity_ordered"]
            description = form.cleaned_data["description"]
            serial_num = form.cleaned_data["serial_num"]
            property_num = form.cleaned_data["property_num"]
            unit_price = form.cleaned_data["unit_price"]

            worksheet[f"B{row}"] = manufacturer
            worksheet[f"C{row}"] = model_part_num
            worksheet[f"D{row}"] = quantity_ordered
            worksheet[f"E{row}"] = description
            worksheet[f"G{row}"] = serial_num
            worksheet[f"H{row}"] = property_num
            worksheet[f"I{row}"] = unit_price
            row += 1

        # Apply custom number format to the last row
        worksheet[f"I{row-1}"].number_format = (
            "_($* #,##0.00_);_($* (#,##0.00);_($* -_0_0_);_(@"
        )
        worksheet[f"J{row-1}"].number_format = (
            "_($* #,##0.00_);_($* (#,##0.00);_($* -_0_0_);_(@"
        )

        # Save the workbook content to the response object
        workbook.save(response)

        return response
