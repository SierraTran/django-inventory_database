from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.styles import Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation


def is_cell_merged(worksheet, cell):
    """
    Checks if a specific cell is merged in the worksheet.

    Args:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object.
        cell (str): The cell reference (e.g., "B31").

    Returns:
        bool: True if the cell is merged, False otherwise.
    """
    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        cell_col, cell_row = range_boundaries(cell + ":" + cell)[0:2]
        if min_col <= cell_col <= max_col and min_row <= cell_row <= max_row:
            return True
    return False


def format_row(worksheet, row):
    """
    Function that formats a new row in the worksheet.

    The function performs the following actions:
    - Applies border formatting to specific cells in the new row.
    - Sets vertical and horizontal alignment for specific cells in the new row.
    - Adds a formula to calculate the total price in a specific cell.

    Args:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object where the row is located.
        row (int): The row index of the new row.
    """
    border_side = Side(style="thin")
    cell_border = Border(
        top=border_side, bottom=border_side, left=border_side, right=border_side
    )
    cell_vert_align = Alignment(vertical="center")
    cell_all_align = Alignment(vertical="center", horizontal="center")

    custom_accounting_style = NamedStyle(
        name="customAccountingStyle",
        number_format="_($* #,##0.00_);_($* (#,##0.00);_($* -_0_0_);_(@",
    )
    if "customAccountingStyle" not in worksheet.parent.named_styles:
        worksheet.parent.add_named_style(custom_accounting_style)

    # Number formatting
    worksheet[f"I{row}"].style = custom_accounting_style
    worksheet[f"J{row}"].style = custom_accounting_style

    # Border formatting
    for letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "M", "N", "O"]:
        worksheet[f"{letter}{row}"].border = cell_border

    # Alignment formatting
    for letter in ["B", "C", "D", "E", "G", "H", "I", "J"]:
        align_format = cell_all_align
        if letter in ["B", "I"]:
            align_format = cell_vert_align
        worksheet[f"{letter}{row}"].alignment = align_format

    # Formula for Total Price
    worksheet[f"J{row}"] = f"=+I{row}*D{row}"
    worksheet[f"J{row+1}"] = f"=+I{row+1}*D{row+1}"
    worksheet[f"J{row+2}"] = f"=+I{row+2}*D{row+2}"


def setup_worksheet(worksheet, itemCount):
    """
    Sets up and adjusts the worksheet to accomodate 8 or more items submitted in the Purchase Order Item Form. 

    Arguments:
        worksheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object that needs to be set up and adjusted.
        itemCount (int): The total number of items that have been submitted.
    """
    amount = itemCount - 7

    # Variables for initial row numbers
    initial_last_item_row = 24
    initial_row_before_notes = 26

    # Variables for new row numbers
    new_last_row = initial_last_item_row + amount
    new_row_before_notes = initial_row_before_notes + amount

    # Remove data validation from cel B31
    dv_to_remove = None
    for dv in worksheet.data_validations.dataValidation:
        if "B31" in dv.sqref:
            dv_to_remove = dv
            break
    if dv_to_remove:
        worksheet.data_validations.dataValidation.remove(dv_to_remove)

    # Unmerge the "Notes" rows
    worksheet.unmerge_cells("C27:H27")
    worksheet.unmerge_cells("B28:H28")
    worksheet.unmerge_cells("B29:H29")
    worksheet.unmerge_cells("B30:H30")
    worksheet.unmerge_cells("B31:H31")

    # Unmerge the "Total" cells
    worksheet.unmerge_cells("I30:I31")
    worksheet.unmerge_cells(start_row=30, start_column=10, end_row=31, end_column=11)

    worksheet.merge_cells("J30:K30")
    worksheet.merge_cells("J31:K31")

    # Add `amount` rows before row 24
    worksheet.insert_rows(initial_last_item_row, amount=amount)

    for row in range(initial_last_item_row, new_last_row + 1):
        # Format cells
        format_row(worksheet, row)

    # Merge new "Notes" cells
    worksheet.merge_cells(f"C{27+amount}:H{27+amount}")  # Shares row with subtotal
    worksheet.merge_cells(f"B{28+amount}:H{28+amount}")  # Shares row with shipping
    worksheet.merge_cells(f"B{29+amount}:H{29+amount}")  # Shares row with tax
    worksheet.merge_cells(f"B{30+amount}:H{30+amount}")  # Shares row with total
    worksheet.merge_cells(f"B{31+amount}:H{31+amount}")  # Shares row with total

    # Data validation object for new last Notes row
    dv = DataValidation(
        type="list", formula1='"ANSI/NCSL Z540-1-1994,ISO/IEC 17025"', allow_blank=True
    )
    new_last_notes_row = f"B{31+amount}"
    worksheet.add_data_validation(dv)
    dv.add(worksheet[new_last_notes_row])

    # Merge new cells for Other accounting cells
    for row in range(27 + amount, 30 + amount):
        if not is_cell_merged(worksheet, f"J{row}"):
            worksheet.merge_cells(f"J{row}:K{row}")

    # Merge new empty description cells
    for row in range(initial_row_before_notes, new_row_before_notes + 1):
        if not is_cell_merged(worksheet, f"E{row}"):
            worksheet.merge_cells(f"E{row}:F{row}")
        worksheet[f"J{row}"].value = None

    # Merge new cells for "Total"
    worksheet.merge_cells(f"I{30+amount}:I{31+amount}")
    start_row = 30 + amount
    end_row = 31 + amount
    worksheet.merge_cells(
        start_row=start_row, start_column=10, end_row=end_row, end_column=11
    )
